import streamlit as st
import pandas as pd
import openpyxl
import urllib.parse
from datetime import datetime
import plotly.express as px
import plotly.graph_objects as go

# ==========================================
# 1. PAGE CONFIG & PREMIUM UI INJECTIONS
# ==========================================
st.set_page_config(page_title="Marine Spares Control Tower", page_icon="🚢", layout="wide")

# Inject Custom CSS for a Premium, Dashboard Aesthetic (No cheap icons)
st.markdown("""
    <style>
    /* Sleek background and font rendering */
    .stApp { background-color: #0E1117; color: #FAFAFA; font-family: 'Inter', sans-serif; }
    
    /* Premium glowing metrics for Critical Alerts */
    .metric-critical {
        background: linear-gradient(135deg, #2b0000 0%, #4a0000 100%);
        border: 1px solid #ff4b4b;
        border-radius: 8px;
        padding: 20px;
        box-shadow: 0 0 15px rgba(255, 75, 75, 0.4);
        text-align: center;
        animation: pulse 2s infinite;
    }
    
    /* Standard sleek metric cards */
    .metric-card {
        background: #1E2329;
        border: 1px solid #2B303A;
        border-radius: 8px;
        padding: 20px;
        text-align: center;
        transition: transform 0.2s ease-in-out;
    }
    .metric-card:hover { transform: translateY(-5px); box-shadow: 0 5px 15px rgba(0,0,0,0.3); }
    
    @keyframes pulse {
        0% { box-shadow: 0 0 10px rgba(255, 75, 75, 0.4); }
        50% { box-shadow: 0 0 20px rgba(255, 75, 75, 0.8); }
        100% { box-shadow: 0 0 10px rgba(255, 75, 75, 0.4); }
    }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# 2. BULLETPROOF DATA EXTRACTION ENGINE
# ==========================================
@st.cache_data(show_spinner=False)
def extract_and_clean_data(file_bytes):
    """100% Extraction Rate Engine: Parses raw XML to bypass Excel display limits."""
    try:
        import io
        file_obj = io.BytesIO(file_bytes)
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        
        # Locate correct sheet dynamically
        sheet_name = next((s for s in wb.sheetnames if 'SPARES' in s.upper()), wb.sheetnames[0])
        ws = wb[sheet_name]
        
        # Dynamic Header Locator
        header_row_idx = next((i + 1 for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)) 
                               if row and "TA REF" in row), None)
        
        if not header_row_idx:
            return pd.DataFrame(), "CRITICAL ERROR: Header row containing 'TA REF' not found."

        headers = [str(cell.value).strip() if cell.value else f"Col_{idx}" for idx, cell in enumerate(ws[header_row_idx])]
        data = []
        base_path = r"Z:\Marine_Dept\Alexis\Spares\Hyperlinks 2026"
        
        # Extract rows
        for row in ws.iter_rows(min_row=header_row_idx + 1):
            row_data = {}
            if not row[headers.index('TA REF') if 'TA REF' in headers else 1].value:
                continue # Skip empty rows instantly
                
            for idx, cell in enumerate(row):
                if idx < len(headers):
                    col_name = headers[idx]
                    row_data[col_name] = cell.value
                    
                    # Absolute Path Rebuilder Algorithm
                    if cell.hyperlink and cell.hyperlink.target:
                        raw_link = urllib.parse.unquote(str(cell.hyperlink.target))
                        if raw_link.startswith("..\\..\\") and "MODION" in raw_link:
                            tail = raw_link.split("MODION")[1].replace("\\", "/")
                            row_data[f"{col_name}_URL"] = f"file:///{base_path.replace(chr(92), '/')}/MODION{tail}"
                        else:
                            row_data[f"{col_name}_URL"] = raw_link
            data.append(row_data)
            
        df = pd.DataFrame(data)
        
        # Ensure mandatory new columns exist to prevent crashes
        if 'SENT TO FINANCE' not in df.columns: df['SENT TO FINANCE'] = pd.NaT
        if 'PRIORITY' not in df.columns: df['PRIORITY'] = "STANDARD"
        
        return df, None
    except Exception as e:
        return pd.DataFrame(), str(e)

# ==========================================
# 3. STATE MACHINE & SLA CALCULATOR
# ==========================================
def process_pipeline_state(df):
    """Calculates exact pipeline state based on strict SLAs."""
    today = pd.to_datetime('today')
    
    # 1. Strict typing for dates
    date_cols = ['DATE', 'SENT TO FINANCE', 'ORDER DATE', 'EST. READINESS', 'RCVD']
    for col in date_cols:
        df[col] = pd.to_datetime(df[col], errors='coerce')
        
    # 2. SLA Targets (Days)
    SLA_SUPPLY = 7
    SLA_FINANCE = 5
    
    statuses = []
    flags = []
    
    for _, row in df.iterrows():
        is_critical = str(row.get('PRIORITY')).upper().strip() == 'CRITICAL'
        
        # Waterfall state logic (End of pipe -> Beginning of pipe)
        if pd.notnull(row['RCVD']):
            state = "🟢 Completed"
            flag = "OK"
        elif pd.notnull(row['EST. READINESS']):
            if row['EST. READINESS'] < today:
                state = "🔴 Logistics Lag (Overdue)"
                flag = "DELAYED"
            else:
                state = "🟡 In Transit"
                flag = "OK"
        elif pd.notnull(row['ORDER DATE']):
            state = "🟠 Ordered (Awaiting Delivery Date)"
            flag = "OK"
        elif pd.notnull(row['SENT TO FINANCE']):
            days_in_finance = (today - row['SENT TO FINANCE']).days
            if days_in_finance > SLA_FINANCE:
                state = "🔴 Finance Lag (Budget Approval Overdue)"
                flag = "DELAYED"
            else:
                state = "🟣 Pending Finance Approval"
                flag = "OK"
        elif pd.notnull(row['DATE']):
            days_in_supply = (today - row['DATE']).days
            if days_in_supply > SLA_SUPPLY:
                state = "🔴 Supply Lag (Quoting Overdue)"
                flag = "DELAYED"
            else:
                state = "🔵 Pending Supply Sourcing"
                flag = "OK"
        else:
            state = "⚪ Unknown"
            flag = "ERROR"
            
        # Priority Override
        if is_critical and flag == "DELAYED":
            state = "🔥 CRITICAL FAILURE: " + state
            
        statuses.append(state)
        flags.append(flag)
        
    df['STATUS'] = statuses
    df['FLAG'] = flags
    
    # Reorder columns logically
    cols = df.columns.tolist()
    front_cols = ['PRIORITY', 'STATUS', 'TA REF']
    front_cols = [c for c in front_cols if c in cols]
    for c in reversed(front_cols):
        cols.insert(0, cols.pop(cols.index(c)))
        
    return df

# ==========================================
# 4. FRONTEND DASHBOARD RENDERING
# ==========================================
st.title("🚢 Marine Spares Control Tower")

uploaded_file = st.file_uploader("Drop Master Spares File (.xlsx) Here", type=["xlsx"])

if uploaded_file:
    with st.spinner("Executing Data Extraction & Path Resolution..."):
        bytes_data = uploaded_file.getvalue()
        raw_df, error = extract_and_clean_data(bytes_data)
        
    if error:
        st.error(f"System Integrity Fault: {error}")
    elif not raw_df.empty:
        df = process_pipeline_state(raw_df)
        
        # --- TOP LEVEL METRICS ---
        st.markdown("<br>", unsafe_allow_html=True)
        col1, col2, col3, col4 = st.columns(4)
        
        total_reqs = len(df)
        completed = len(df[df['STATUS'] == '🟢 Completed'])
        critical_alerts = len(df[df['STATUS'].str.contains("CRITICAL FAILURE")])
        total_delayed = len(df[df['FLAG'] == 'DELAYED'])
        
        col1.markdown(f"<div class='metric-card'><h3>{total_reqs}</h3><p>Total Pipeline Requisitions</p></div>", unsafe_allow_html=True)
        col2.markdown(f"<div class='metric-card'><h3>{completed}</h3><p>Completed / Onboard</p></div>", unsafe_allow_html=True)
        col3.markdown(f"<div class='metric-card'><h3>{total_delayed - critical_alerts}</h3><p>Standard Delays</p></div>", unsafe_allow_html=True)
        
        if critical_alerts > 0:
            col4.markdown(f"<div class='metric-critical'><h3 style='color:white;'>{critical_alerts}</h3><p style='color:white; font-weight:bold;'>CRITICAL OVERDUE</p></div>", unsafe_allow_html=True)
        else:
            col4.markdown(f"<div class='metric-card'><h3>0</h3><p>Critical Overdue</p></div>", unsafe_allow_html=True)

        st.markdown("<hr style='border:1px solid #2B303A'>", unsafe_allow_html=True)
        
        # --- THE BOTTLENECK X-RAY (PLOTLY ANIMATIONS) ---
        st.subheader("📊 Pipeline Bottleneck Analysis")
        
        # Filter out completed items to see the active pipeline
        active_df = df[df['STATUS'] != '🟢 Completed']
        
        if not active_df.empty:
            # Create a premium Donut chart
            status_counts = active_df['STATUS'].value_counts().reset_index()
            status_counts.columns = ['Status', 'Count']
            
            fig = px.pie(
                status_counts, 
                values='Count', 
                names='Status', 
                hole=0.6,
                color_discrete_sequence=px.colors.qualitative.Pastel
            )
            fig.update_layout(
                plot_bgcolor='rgba(0,0,0,0)', 
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color='#FAFAFA'),
                margin=dict(t=20, b=20, l=20, r=20),
                showlegend=True
            )
            fig.update_traces(textposition='inside', textinfo='percent+label', hoverinfo='label+value')
            
            col_chart, col_triage = st.columns([1.5, 2])
            
            with col_chart:
                st.plotly_chart(fig, use_container_width=True)
                
            with col_triage:
                st.markdown("### 🔥 Priority Triage Zone")
                critical_df = df[df['STATUS'].str.contains("CRITICAL FAILURE")]
                if not critical_df.empty:
                    st.error("Immediate Action Required on the following items:")
                    display_cols = ['TA REF', 'EQUIPMENT', 'STATUS', 'DATE', 'SENT TO FINANCE']
                    display_cols = [c for c in display_cols if c in critical_df.columns]
                    st.dataframe(critical_df[display_cols], hide_index=True, use_container_width=True)
                else:
                    st.success("No critical overdue items. Fleet equipment is secure.")
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        # --- INTERACTIVE DATA GRID ---
        st.subheader("🔍 Full Fleet Control Grid")
        
        # Advanced Filtering
        f_col1, f_col2, f_col3 = st.columns(3)
        with f_col1:
            stat_filt = st.multiselect("Filter Status:", options=df['STATUS'].unique(), default=df['STATUS'].unique())
        with f_col2:
            equip_filt = st.multiselect("Filter Equipment:", options=df['EQUIPMENT'].dropna().unique())
        with f_col3:
            pri_filt = st.multiselect("Priority:", options=df['PRIORITY'].unique(), default=df['PRIORITY'].unique())
            
        filtered = df[df['STATUS'].isin(stat_filt) & df['PRIORITY'].isin(pri_filt)]
        if equip_filt:
            filtered = filtered[filtered['EQUIPMENT'].isin(equip_filt)]
            
        # Premium Column Config
        st.dataframe(
            filtered,
            use_container_width=True,
            hide_index=True,
            column_config={
                "TA REF_URL": st.column_config.LinkColumn("Document Link", display_text="Open File"),
                "COST": st.column_config.NumberColumn("Cost ($)", format="$%.2f"),
                "DATE": st.column_config.DateColumn("Requested", format="DD MMM YYYY"),
                "SENT TO FINANCE": st.column_config.DateColumn("To Finance", format="DD MMM YYYY"),
                "ORDER DATE": st.column_config.DateColumn("PO Placed", format="DD MMM YYYY")
            }
        )
