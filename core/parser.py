"""
core/parser.py — Marine Spares Data Engine  v5
Surgical fixes vs v4:

ROW CLASSIFICATION (definitive):
  ┌─────────────────────────────────────────────────────────────────────┐
  │ Type          │ NR    │ SEQ   │ TA REF  │ Action                   │
  │ Normal        │ any   │ any   │ present │ Full record              │
  │ Continuation  │ None  │ None  │ present │ Record + inherit parent  │
  │ Ghost/Split   │ None  │ None  │ absent  │ Merge into parent        │
  │ Blank         │ –     │ –     │ –       │ Skip                     │
  └─────────────────────────────────────────────────────────────────────┘

CONTINUATION ROW (OF-26-2569 pattern in STEFANOS):
  - Has its own TA REF but no NR / SEQ
  - Inherits case_code + account_code from immediately preceding record
  - Keeps its own data (description, supplier, dates, cost, hyperlink)

GHOST/SPLIT ROW:
  - No TA REF, no NR, no SEQ — but has supplier/order_date/cost
  - Merged into previous record as sub_order
  - Cost NOT merged if parent is cancelled

_norm_code: accepts any Python type (int/float/str/None) — no .strip() on raw value

invoice: coerced with format='mixed' to handle '60 DAYS' strings silently
"""
from __future__ import annotations

import io
import re
import urllib.parse
from datetime import datetime, date
from typing import Optional

import openpyxl
import pandas as pd

# ── Column mapping ─────────────────────────────────────────────────────────────
COL_MAP: dict[str, str] = {
    "NR":             "nr",
    "TA REF":         "ta_ref",
    "CASE":           "case_code",
    " ":              "seq",
    "REF":            "ref_date",
    "EQUIPMENT":      "equipment",
    "DESCRIPTION":    "description",
    "DATE":           "date_requested",
    "MESSAGE":        "message",
    "ORDERED":        "supplier",
    "CODE":           "account_code",
    "CONFIRMATION ":  "confirmation",
    "ORDER DATE":     "order_date",
    "COST":           "cost",
    "EST. READINESS": "est_readiness",
    "PORT ":          "port",
    "AWB":            "awb",
    "RCVD":           "rcvd",
    "INVOICE":        "invoice",
}

SLA: dict[str, int] = {
    "supply":  7,
    "finance": 5,
    "ordered": 45,
    "transit": 21,
}

_HYPERLINK_BASE = r"Z:\Marine_Dept\Alexis\Spares\Hyperlinks 2026"

_PRIORITY_COLS = [
    "vessel", "status_label", "flag",
    "ta_ref", "nr", "seq", "case_code", "description", "equipment",
    "category_name", "date_requested", "supplier", "order_date",
    "cost", "cost_raw", "is_cancelled",
    "est_readiness", "port", "rcvd", "account_code", "message",
    "confirmation", "awb", "invoice", "document_url",
    "status", "sla_breach", "sla_days_over", "days_in_stage", "sub_orders",
]

# Regex: embedded relative path in cell value
_PATH_RE = re.compile(
    r"(\.\.[\\/][^\s\x00]+\.(?:doc|pdf|xls|xlsx|docx))", re.IGNORECASE
)


# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def _resolve_hyperlink(raw: str) -> str:
    decoded = urllib.parse.unquote(str(raw))
    m = re.search(r"MODION(.+)$", decoded, re.IGNORECASE)
    if m:
        tail = m.group(1).replace("\\", "/")
        base = _HYPERLINK_BASE.replace("\\", "/")
        return f"file:///{base}/MODION{tail}"
    return decoded


def _extract_path_from_value(val: object) -> Optional[str]:
    """Extract embedded path string from a cell value (fallback for cells
    where hyperlink target is missing but value IS the path)."""
    if val is None:
        return None
    m = _PATH_RE.search(str(val))
    return _resolve_hyperlink(m.group(1)) if m else None


def _ts(val) -> Optional[pd.Timestamp]:
    if val is None:
        return None
    try:
        r = pd.Timestamp(val)
        return None if pd.isnull(r) else r
    except Exception:
        return None


def _is_cancelled(confirmation: object) -> bool:
    return "CANCEL" in str(confirmation or "").upper()


def _norm_code(x: object) -> str:
    """
    Normalise any account_code value to a clean string key.
    Accepts int (5513), float (5513.0), str ('5517C'), None — no type assumption.
    """
    if x is None:
        return ""
    s = str(x).strip()
    if s in ("", "None", "nan", "NaN", "<NA>"):
        return ""
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
        return s
    except (ValueError, OverflowError):
        return s


def _doc_url(cell_hl_target: Optional[str], cell_value: object) -> Optional[str]:
    """Resolve document URL from hyperlink target OR embedded path in value."""
    if cell_hl_target:
        return _resolve_hyperlink(str(cell_hl_target))
    return _extract_path_from_value(cell_value)


# ──────────────────────────────────────────────────────────────────────────────
# STATE MACHINE
# ──────────────────────────────────────────────────────────────────────────────

def _compute_state(row: dict, now: pd.Timestamp) -> dict:
    rcvd       = _ts(row.get("rcvd"))
    est_ready  = _ts(row.get("est_readiness"))
    order_date = _ts(row.get("order_date"))
    date_req   = _ts(row.get("date_requested"))
    confirm    = str(row.get("confirmation") or "").upper()

    if "CANCEL" in confirm:
        return dict(status="CANCELLED", status_label="✖ Cancelled",
                    flag="CANCELLED", days_in_stage=None,
                    sla_breach=False, sla_days_over=0)

    if rcvd is not None:
        return dict(status="RECEIVED", status_label="🟢 Received",
                    flag="OK", days_in_stage=int((now - rcvd).days),
                    sla_breach=False, sla_days_over=0)

    if est_ready is not None:
        overdue = int((now - est_ready).days)
        breach  = overdue > 0
        return dict(
            status        = "OVERDUE_TRANSIT" if breach else "IN_TRANSIT",
            status_label  = "🔴 Transit Overdue" if breach else "🟡 In Transit",
            flag          = "DELAYED" if breach else "OK",
            days_in_stage = abs(overdue),
            sla_breach    = breach,
            sla_days_over = overdue if breach else 0,
        )

    if order_date is not None:
        days  = int((now - order_date).days)
        breach = days > SLA["ordered"]
        return dict(
            status        = "OVERDUE_ORDERED" if breach else "ORDERED",
            status_label  = "🔴 Order Overdue" if breach else "🟠 Ordered",
            flag          = "DELAYED" if breach else "OK",
            days_in_stage = days,
            sla_breach    = breach,
            sla_days_over = max(0, days - SLA["ordered"]),
        )

    if date_req is not None:
        days  = int((now - date_req).days)
        breach = days > SLA["supply"]
        return dict(
            status        = "OVERDUE_SUPPLY" if breach else "PENDING_SUPPLY",
            status_label  = "🔴 Supply Overdue" if breach else "🔵 Pending Supply",
            flag          = "DELAYED" if breach else "OK",
            days_in_stage = days,
            sla_breach    = breach,
            sla_days_over = max(0, days - SLA["supply"]),
        )

    return dict(status="UNKNOWN", status_label="⚪ Unknown",
                flag="ERROR", days_in_stage=None,
                sla_breach=False, sla_days_over=0)


# ──────────────────────────────────────────────────────────────────────────────
# INDEX SHEET PARSER
# ──────────────────────────────────────────────────────────────────────────────

def _parse_index(ws) -> dict[str, dict]:
    result: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        padded = (list(row) + [None] * 10)[:10]
        name, case, code = padded[0], padded[1], padded[2]
        cost, rcvd_count, processed = padded[5], padded[7], padded[9]
        if code is None:
            continue
        code_str = _norm_code(code)
        if not code_str:
            continue
        result[code_str] = {
            "category_name":          str(name).strip() if name else "",
            "case_code":              str(case).strip() if case else "",
            "cost":                   float(cost) if cost else 0.0,
            "requisitions_received":  int(rcvd_count) if rcvd_count else 0,
            "requisitions_processed": int(processed) if processed else 0,
        }
    return result


def _extract_vessel(ws) -> str:
    cell = ws.cell(row=1, column=1).value
    if not cell:
        return ""
    s = str(cell)
    m = re.match(r"M/V\s+(.+?)\s*[-–]", s, re.IGNORECASE)
    return m.group(1).strip().title() if m else s.strip()


# ──────────────────────────────────────────────────────────────────────────────
# ROW CLASSIFIER
# ──────────────────────────────────────────────────────────────────────────────

def _classify_row(raw: dict) -> str:
    """
    Returns: 'normal' | 'continuation' | 'ghost' | 'blank'

    normal:       ta_ref present                          → full record
    continuation: ta_ref present, nr=None, seq=None       → record + inherit parent
    ghost:        ta_ref absent + has supplier/date/cost   → merge into parent
    blank:        nothing at all                           → skip
    """
    ta_ref = raw.get("ta_ref")
    nr     = raw.get("nr")
    seq    = raw.get("seq")
    has_payload = (raw.get("supplier") is not None
                   or raw.get("order_date") is not None
                   or raw.get("cost") is not None)

    data_vals = [v for k, v in raw.items()
                 if not k.startswith("__hl_") and v is not None]

    if not data_vals:
        return "blank"

    if ta_ref is not None:
        # Has a TA REF → always a real requisition
        if nr is None and seq is None:
            return "continuation"   # e.g. OF-26-2569 in STEFANOS
        return "normal"

    # No TA REF
    if has_payload:
        return "ghost"              # e.g. OCEANTECH sub-order row
    return "blank"


# ──────────────────────────────────────────────────────────────────────────────
# MAIN PARSER
# ──────────────────────────────────────────────────────────────────────────────

def parse_workbook(file_bytes: bytes) -> tuple[pd.DataFrame, dict, list[str]]:
    """
    Parse any Marine Spares workbook.

    Row classification:
      normal       → full record (NR/SEQ/TA REF all present)
      continuation → TA REF present, NR+SEQ absent; inherits case_code + account_code
                     from preceding record; retains its own data (dates, costs, etc.)
      ghost/split  → TA REF absent; merged into preceding record as sub_order
      blank        → skipped

    Financial guarantee:
      Cancelled rows: cost set to NaN (excluded from all sums)
                      cost_raw preserved for audit
      Ghost cost: only merged if parent is NOT cancelled
    """
    warnings_out: list[str] = []
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    spares_name = next((s for s in wb.sheetnames if "SPARES" in s.upper()), None)
    index_name  = next((s for s in wb.sheetnames if "INDEX"  in s.upper()), None)
    if not spares_name:
        raise ValueError("Cannot find a SPARES sheet in this workbook.")

    ws         = wb[spares_name]
    index_kpis = _parse_index(wb[index_name]) if index_name else {}
    vessel     = _extract_vessel(ws)

    # ── Find header row ───────────────────────────────────────────────────────
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
        if row and "TA REF" in row:
            header_row = i + 1
            break
    if header_row is None:
        raise ValueError("Cannot find header row containing 'TA REF'.")

    col_idx: dict[int, str] = {}
    for i, cell in enumerate(ws[header_row]):
        if cell.value is not None:
            raw_h = str(cell.value)
            col_idx[i] = COL_MAP.get(raw_h,
                                     raw_h.lower().replace(" ", "_").strip())

    # ── Read and classify rows ────────────────────────────────────────────────
    records: list[dict] = []

    for row in ws.iter_rows(min_row=header_row + 1):
        raw: dict = {}
        for i, cell in enumerate(row):
            if i not in col_idx:
                continue
            raw[col_idx[i]] = cell.value
            if cell.hyperlink and cell.hyperlink.target:
                raw[f"__hl_{col_idx[i]}"] = cell.hyperlink.target

        kind = _classify_row(raw)

        # ── BLANK ─────────────────────────────────────────────────────────────
        if kind == "blank":
            continue

        # ── GHOST / SPLIT ORDER ───────────────────────────────────────────────
        if kind == "ghost":
            if records:
                parent   = records[-1]
                sub_cost = raw.get("cost")
                parent.setdefault("sub_orders", []).append({
                    "supplier":   raw.get("supplier"),
                    "order_date": raw.get("order_date"),
                    "cost":       sub_cost,
                })
                # Accumulate cost only if parent is NOT cancelled
                if sub_cost is not None and not parent.get("_cancelled_flag", False):
                    parent["cost"] = (parent.get("cost") or 0.0) + float(sub_cost)
                warnings_out.append(
                    f"Split order merged → {parent.get('ta_ref','?')} "
                    f"(supplier: {raw.get('supplier')}, "
                    f"cost: {sub_cost})"
                )
            else:
                warnings_out.append(f"Orphan ghost row ignored: {raw}")
            continue

        # ── CONTINUATION ROW (TA REF present, NR+SEQ absent) ─────────────────
        if kind == "continuation":
            # Inherit case_code + account_code from the most recent record
            # (they share the same requisition group / vessel area)
            if records:
                parent_rec = records[-1]
                if not raw.get("case_code"):
                    raw["case_code"]    = parent_rec.get("case_code")
                if not raw.get("account_code"):
                    raw["account_code"] = parent_rec.get("account_code")
            warnings_out.append(
                f"Continuation row kept: {raw.get('ta_ref')} "
                f"(NR/SEQ absent, inherited case/code from preceding record)"
            )
            # falls through to normal record processing below

        # ── NORMAL / CONTINUATION — build record ──────────────────────────────
        # Document URL: hyperlink object on MESSAGE takes priority;
        # fall back to embedded path string in MESSAGE cell value
        msg_hl  = raw.pop("__hl_message", None)
        msg_val = raw.get("message")
        raw["document_url"] = _doc_url(msg_hl, msg_val)

        # Remove remaining __hl_ keys
        for k in [k for k in list(raw) if k.startswith("__hl_")]:
            raw.pop(k)

        raw["_cancelled_flag"] = _is_cancelled(raw.get("confirmation"))
        raw["vessel"]          = vessel
        raw.setdefault("sub_orders", [])
        records.append(raw)

    if not records:
        raise ValueError("No valid requisition rows found.")

    df = pd.DataFrame(records)

    # ── Date coercion ──────────────────────────────────────────────────────────
    # format='mixed' silently handles text like '60 DAYS', 'ARRIVING ...' → NaT
    for col in ["date_requested", "order_date", "est_readiness",
                "rcvd", "ref_date", "invoice"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce", format="mixed")

    # ── Numeric coercion ───────────────────────────────────────────────────────
    for col in ["cost", "seq", "nr"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # ── String coercion ────────────────────────────────────────────────────────
    str_cols = ["ta_ref", "case_code", "equipment", "description", "message",
                "supplier", "confirmation", "port", "awb", "vessel"]
    for col in str_cols:
        if col in df.columns:
            df[col] = (df[col].astype(str)
                               .str.strip()
                               .replace({"None": "", "nan": "", "NaN": ""}))

    # ── account_code normalisation ─────────────────────────────────────────────
    # _norm_code accepts any type — no astype(str) needed first
    if "account_code" in df.columns:
        df["account_code"] = df["account_code"].apply(_norm_code)
    else:
        df["account_code"] = ""

    # ── Cancelled cost isolation ───────────────────────────────────────────────
    df["is_cancelled"] = df["_cancelled_flag"].fillna(False).astype(bool)
    df["cost_raw"]     = pd.to_numeric(df.get("cost"), errors="coerce").copy()
    df.loc[df["is_cancelled"], "cost"] = float("nan")
    df = df.drop(columns=["_cancelled_flag"])

    # ── Category enrichment from INDEX ────────────────────────────────────────
    df["category_name"] = df["account_code"].apply(
        lambda c: index_kpis.get(c, {}).get("category_name", "")
    )

    # ── State machine ──────────────────────────────────────────────────────────
    now    = pd.Timestamp(datetime.now())
    states = [_compute_state(r, now) for r in df.to_dict("records")]
    df     = pd.concat([df, pd.DataFrame(states)], axis=1)

    # ── Column ordering ────────────────────────────────────────────────────────
    ordered = [c for c in _PRIORITY_COLS if c in df.columns]
    rest    = [c for c in df.columns    if c not in ordered]
    return df[ordered + rest], index_kpis, warnings_out
